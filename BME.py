import openpyxl

def search_name_in_excel(file_path, sheet_name, target_name):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]

        total_rows = 50  # 1 for header + 49 rows of data
        total_columns = 17 # Update this if you have more/less columns

        # Read headers from row 1 (A1 to AW1 for 49 columns)
        headers = [sheet.cell(row=1, column=col).value for col in range(1, total_columns + 1)]

        # Loop through rows 2 to 50 (49 names starting from row 2)
        found = False
        for row in range(1, total_rows + 1):
            name_in_cell = sheet.cell(row=row, column=3).value  # Column C = 3
            if str(name_in_cell).strip().lower() == target_name.strip().lower():
                found = True
                data = [sheet.cell(row=row, column=col).value for col in range(1, total_columns + 1)]

                print("\n📄Entire Details:\n")
                for h, d in zip(headers, data):
                    if h and str(h).strip() :
                        print(f"{str(h).capitalize()}: {d}")
                    else:
                        print(f"(No Header): {d}")
                break

        if not found:
            print(f"❌ Name '{target_name}' not found in column C.")

        wb.close()

    except Exception as e:
        print("❌ Error:", e)

# === Example usage ===

file_path = r"E:/MEC/IT AND BME/BME DETAILS.xlsx"  # ✅ Update with your actual path
sheet_name = "FINAL"

# 👤 Take user input for the name
user_input = input("Enter the name to search: ")

# 🔍 Run the search
search_name_in_excel(file_path, sheet_name, user_input)

        
